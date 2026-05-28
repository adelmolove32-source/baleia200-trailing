import pandas as pd, numpy as np, time, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

exec(open(r'C:\Users\muril\Desktop\bot btc2\baleia200_pptx.py').read().split('# =====================================================')[0])

DATA = r'C:\Users\muril\Desktop\bot btc2\btc_5m_360d.pkl'
df = pd.read_pickle(DATA)
cutoff = df['timestamp'].max() - pd.Timedelta(days=360)
bt = df[df['timestamp']>=cutoff].copy().reset_index(drop=True)
bt = build_signals(bt, sma_shift=1)
bt = bt.dropna(subset=['sma20','sma200']).reset_index(drop=True)

t0 = time.time()
trades = run(bt, 'midpoint_trail', 0.0, 5)
print(f'Backtest: {time.time()-t0:.1f}s, {len(trades)} trades')

cap = 200.0
bal = cap
results = []
eq = [cap]
for t in trades:
    bal += t['pnl']
    t['banca'] = bal
    eq.append(bal)
    stop_dist = abs(t['entry']-t['stop'])
    risco_pct = stop_dist/t['entry']*100
    r_val = t['pnl']/(stop_dist/t['entry']*t['alloc']) if stop_dist>0 else 0
    resultado = 'WIN' if t['pnl']>0 else 'LOSS'
    motivo = t['rz']
    if motivo == 'trail_win':
        motivo = 'TRAIL'
    elif motivo == 'sma200_reversal':
        if t['pnl'] > 0:
            motivo = 'REVERSAL'
        else:
            motivo = 'REVERSAL'
    elif motivo == 'alvo':
        motivo = 'TARGET'
    elif motivo == 'stop':
        motivo = 'STOP'
    results.append({
        'entrada_data': pd.Timestamp(t['entry_ts']).strftime('%d/%m/%Y'),
        'entrada_hora': pd.Timestamp(t['entry_ts']).strftime('%H:%M'),
        'saida_data': pd.Timestamp(t['exit_ts']).strftime('%d/%m/%Y'),
        'saida_hora': pd.Timestamp(t['exit_ts']).strftime('%H:%M'),
        'direcao': t['tipo'],
        'entrada': t['entry'],
        'stop': t['stop'],
        'alvo': t['target'],
        'risco_pct': round(risco_pct, 3),
        'r': round(r_val, 2),
        'pnl': round(t['pnl'], 2),
        'banca': round(bal, 2),
        'resultado': resultado,
        'motivo': motivo,
        'sinal': t.get('sinal', '')
    })

eq = np.array(eq)
peak = np.maximum.accumulate(eq)
dd = (peak-eq)/peak*100
max_dd = dd.max()

wins = [r for r in results if r['resultado']=='WIN']
losses = [r for r in results if r['resultado']=='LOSS']
wr = len(wins)/len(results)*100

pnls = np.array([r['pnl'] for r in results])
maior_win = pnls.max()
maior_loss = pnls.min()

longs = [r for r in results if r['direcao']=='LONG']
shorts = [r for r in results if r['direcao']=='SHORT']
long_wins = [r for r in longs if r['resultado']=='WIN']
short_wins = [r for r in shorts if r['resultado']=='WIN']
long_pnl = sum(r['pnl'] for r in longs)
short_pnl = sum(r['pnl'] for r in shorts)

saidas = {}
for r in results:
    m = r['motivo']
    if m not in saidas:
        saidas[m] = {'n':0,'w':0,'pnl':0}
    saidas[m]['n'] += 1
    if r['resultado']=='WIN':
        saidas[m]['w'] += 1
    saidas[m]['pnl'] += r['pnl']

# ====== CREATE WORKBOOK ======
wb = Workbook()

# ===== STYLES =====
title_font = Font(name='Arial', bold=True, size=16)
subtitle_font = Font(name='Arial', bold=True, size=11)
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
green_font = Font(name='Arial', size=10, color='006100')
red_font = Font(name='Arial', size=10, color='9C0006')

# ===== SHEET 1: Dashboard =====
ws1 = wb.active
ws1.title = 'Dashboard'
ws1.sheet_properties.tabColor = '2F5496'

ws1.merge_cells('A1:F1')
ws1['A1'] = 'BALEIA 200 - TRAILING STOP'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A2:F2')
ws1['A2'] = 'BTCUSDT | 5m | 360 dias | jun/2025 a mai/2026'
ws1['A2'].font = subtitle_font
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Resumo
row = 4
ws1.merge_cells(f'A{row}:F{row}')
ws1[f'A{row}'] = 'RESUMO'
ws1[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

stats = [
    ('Capital Inicial', '$ 200.00'),
    ('Capital Final', f'$ {eq[-1]:.2f}'),
    ('Retorno Total', f'+{((eq[-1]/200-1)*100):.2f}%'),
    ('Drawdown Maximo', f'{max_dd:.2f}%'),
    ('Total Trades', str(len(results))),
    ('Vitorias', f'{len(wins)} ({wr:.1f}%)'),
    ('Derrotas', f'{len(losses)} ({(100-wr):.1f}%)'),
    ('Maior Win', f'$ {maior_win:.2f}'),
    ('Maior Loss', f'$ {maior_loss:.2f}'),
    ('PnL Medio', f'$ {pnls.mean():.2f}'),
    ('Mediana PnL', f'$ {np.median(pnls):.2f}'),
]
for i, (k, v) in enumerate(stats):
    r = row + 1 + i
    ws1[f'A{r}'] = k
    ws1[f'A{r}'].font = Font(name='Arial', bold=True, size=10)
    ws1[f'A{r}'].alignment = left_align
    ws1[f'B{r}'] = v
    ws1[f'B{r}'].font = data_font
    ws1[f'B{r}'].alignment = left_align

# Distribuicao
row = 16
ws1.merge_cells(f'A{row}:F{row}')
ws1[f'A{row}'] = 'DISTRIBUICAO'
ws1[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

row += 1
headers = ['Direcao', 'Trades', 'Wins', 'Losses', 'WR %', 'PnL $']
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

for i, (label, n, w, l, wr_pct, pnl) in enumerate([
    ('LONG', len(longs), len(long_wins), len(longs)-len(long_wins),
     len(long_wins)/len(longs)*100 if longs else 0, long_pnl),
    ('SHORT', len(shorts), len(short_wins), len(shorts)-len(short_wins),
     len(short_wins)/len(shorts)*100 if shorts else 0, short_pnl),
]):
    r = row + 1 + i
    for c, v in enumerate([label, n, w, l, f'{wr_pct:.1f}%', f'$ {pnl:.2f}'], 1):
        cell = ws1.cell(row=r, column=c, value=v)
        cell.font = data_font
        cell.alignment = center
        cell.border = thin_border

# Motivo saida
row = 20
ws1.merge_cells(f'A{row}:E{row}')
ws1[f'A{row}'] = 'MOTIVO SAIDA'
ws1[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

row += 1
headers2 = ['Motivo', 'Trades', 'Wins', 'WR %', 'PnL $']
for c, h in enumerate(headers2, 1):
    cell = ws1.cell(row=row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

row += 1
for m in ['TARGET', 'TRAIL', 'REVERSAL', 'STOP']:
    if m in saidas:
        v = saidas[m]
        wr_s = v['w']/v['n']*100
        for c, val in enumerate([m, v['n'], v['w'], f'{wr_s:.0f}%', f'$ {v["pnl"]:.2f}'], 1):
            cell = ws1.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

# Ajustar larguras
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14

# ===== SHEET 2: Operacoes =====
ws2 = wb.create_sheet('Operacoes')
ws2.sheet_properties.tabColor = '548235'

ws2.merge_cells('A1:N1')
ws2['A1'] = 'OPERACOES - BALEIA 200 TRAILING STOP'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')

headers3 = ['Data Entrada', 'Hora Entrada', 'Data Saida', 'Hora Saida',
            'Direcao', 'Entrada', 'Stop', 'Alvo', 'Risco %', 'R', 'PnL $',
            'Banca $', 'Resultado', 'Motivo Saida', 'Sinal']
for c, h in enumerate(headers3, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

for i, r in enumerate(results):
    row = i + 3
    vals = [r['entrada_data'], r['entrada_hora'], r['saida_data'], r['saida_hora'],
            r['direcao'], r['entrada'], r['stop'], r['alvo'], r['risco_pct'],
            r['r'], r['pnl'], r['banca'], r['resultado'], r['motivo'], r['sinal']]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.font = data_font
        cell.alignment = center
        cell.border = thin_border
        if r['resultado'] == 'WIN':
            cell.font = green_font
        elif r['resultado'] == 'LOSS':
            cell.font = red_font
    # Formata numeros
    ws2.cell(row=row, column=6).number_format = '#,##0.00'
    ws2.cell(row=row, column=7).number_format = '#,##0.00'
    ws2.cell(row=row, column=8).number_format = '#,##0.00'
    ws2.cell(row=row, column=9).number_format = '0.000'
    ws2.cell(row=row, column=10).number_format = '0.00'
    ws2.cell(row=row, column=11).number_format = '0.00'
    ws2.cell(row=row, column=12).number_format = '0.00'

# Larguras Operacoes
widths = [16, 14, 16, 14, 10, 12, 12, 12, 10, 8, 10, 10, 10, 14, 10]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Congelar painel
ws2.freeze_panes = 'A3'

# ===== SHEET 3: Evolucao Banca =====
ws3 = wb.create_sheet('Evolucao Banca')
ws3.sheet_properties.tabColor = 'BF8F00'

ws3.merge_cells('A1:B1')
ws3['A1'] = 'EVOLUCAO BANCA - BALEIA 200 TRAILING STOP'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws3['A2'] = 'Trade'
ws3['B2'] = 'Banca $'
for c in [1, 2]:
    cell = ws3.cell(row=2, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

for i, e in enumerate(eq):
    row = i + 3
    ws3.cell(row=row, column=1, value=i).font = data_font
    ws3.cell(row=row, column=1).alignment = center
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=2, value=round(e, 2)).font = data_font
    ws3.cell(row=row, column=2).alignment = center
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).number_format = '0.00'

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 14
ws3.freeze_panes = 'A3'

# Salvar
path = r'C:\Users\muril\Desktop\bot btc2\baleia200_trailing_360d.xlsx'
wb.save(path)
print(f'\nPlanilha salva: {path}')
print(f'  Dashboard: resumo + {len(results)} operacoes')
print(f'  Equity: {len(eq)} pontos')
